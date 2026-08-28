"""Massen-Import von Preislisten aus Excel (.xlsx) und PDF.

Erkennt Spalten (Bezeichnung/Einheit/EP/Stundensatz/BKP) tolerant und
haengt sie an die eigene Preisliste (data_store PREISE_PATH) an.

Benoetigt: openpyxl (Excel), pdfplumber (PDF) im Bundle-Runtime.
"""
import csv
import os
import re

from . import data_store as ds
from . import firmen_preise as fp


def _norm(s):
    return (s or "").strip().lower()


def _to_float(v):
    if v is None:
        return None
    try:
        return float(str(v).replace("'", "").replace(" ", "").replace(",", "."))
    except (ValueError, TypeError):
        return None


def _parse_zeile(line: str) -> dict:
    """Zerlegt eine leerzeichen-getrennte Preislisten-Zeile.

    Formate: '1.1 Innenanstrich Wand  m2  42.50'  oder
              'Bodenbelag verlegen   m2   25.0'
    Letzte Zahl = EP, vorletztes Wort = Einheit, Rest = Bezeichnung.
    """
    line = line.strip()
    # bkp-praefix? (z.b. '1.1 ')
    m = re.match(r"^(\d+(\.\d+)*)\s+(.*)$", line)
    bkp = ""
    if m:
        bkp = m.group(1)
        rest = m.group(3)
    else:
        rest = line
    parts = rest.rsplit(None, 2)  # [..., einheit, betrag] oder [bez, einheit, betrag]
    if len(parts) >= 3:
        bez, einh, ep = parts[0], parts[-2], parts[-1]
        # bez kann noch aus mehreren woertern bestehen -> alles vor einheit
        bez = rest[:rest.rfind(einh)].strip()
    elif len(parts) == 2:
        bez, ep = parts[0], parts[-1]
        einh = ""
    else:
        bez, ep, einh = rest, "", ""
    ep_v = _to_float(ep)
    return {"bkp": bkp, "bezeichnung": bez.strip(), "einheit": einh.strip(),
            "ep_chf": ("" if ep_v is None else ep_v), "stundensatz_chf": "",
            "kosten_chf": "", "kategorie": ""}


def _spalten_merken(zeile_dict):
    """Mappt eine gelesene Zeile auf unser Schema."""
    d = {_norm(k): v for k, v in zeile_dict.items() if k}
    bez = (d.get("bezeichnung") or d.get("text") or d.get("artikel")
           or d.get("bezeichnung / text") or d.get("leistung") or "").strip()
    einh = (d.get("einheit") or d.get("me") or "").strip()
    ep = _to_float(d.get("ep_chf") or d.get("ep") or d.get("einheitspreis")
                   or d.get("chf") or d.get("preis") or "")
    std = _to_float(d.get("stundensatz_chf") or d.get("stundensatz") or "")
    bkp = (d.get("bkp") or d.get("npk") or d.get("kapitel") or d.get("nummer") or "").strip()
    kat = (d.get("kategorie") or d.get("gewerk") or "").strip()
    return {
        "bkp": bkp, "bezeichnung": bez, "einheit": einh,
        "ep_chf": ("" if ep is None else ep),
        "stundensatz_chf": ("" if std is None else std),
        "kosten_chf": "", "kategorie": kat,
    }


# --- Excel ----------------------------------------------------------------
def import_xlsx(pfad: str) -> int:
    """Liest alle Blaetter einer .xlsx und gibt Anzahl uebernommener Preise zurueck."""
    from openpyxl import load_workbook
    wb = load_workbook(pfad, data_only=True, read_only=True)
    n = 0
    for ws in wb.worksheets:
        rows = ws.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            continue
        # spaltenueberschriften normalisieren
        cols = {_norm(str(h)): i for i, h in enumerate(header) if h}
        for r in rows:
            if not r or all(c is None for c in r):
                continue
            zeile = {str(k): (r[i] if i < len(r) else None) for k, i in cols.items()}
            z = _spalten_merken(zeile)
            if not z["bezeichnung"]:
                continue
            fp.zeile_speichern(z)
            n += 1
    return n


# --- PDF ------------------------------------------------------------------
def import_pdf(pfad: str) -> int:
    """Liest Tabellen aus einer PDF-Preisliste (pypdf) und uebernimmt sie.

    pypdf ist im Bundle signiert (kein externes charset_normalizer noetig).
    """
    from pypdf import PdfReader
    n = 0
    reader = PdfReader(pfad)
    for page in reader.pages:
        try:
            tbls = page.extract_table()
        except Exception:
            tbls = None
        if not tbls:
            # Fallback: rohen Text zeilenweise parsen, spalten aus leerzeichen-getrenntem layout
            text = page.extract_text() or ""
            tbls = []
            for ln in text.splitlines():
                ln = ln.strip()
                if not ln:
                    continue
                # NUR echte positions-zeilen: muessen mit pos-nummer beginnen (z.b. '5' oder '1.1')
                if not re.match(r"^\d+(\.\d+)*\s", ln):
                    continue
                tbls.append([ln])
        for tbl in (tbls if isinstance(tbls, list) else [tbls]):
            if not tbl:
                continue
            # tbl kann liste von reihen sein oder einzelne zeile
            rows = tbl if isinstance(tbl, list) else [tbl]
            if not rows:
                continue
            header = rows[0]
            if isinstance(header, str):
                # einzelzeile (leerzeichen-getrenntes layout) -> spalten extrahieren
                z = _parse_zeile(header)
                if z and z["bezeichnung"]:
                    fp.zeile_speichern(z)
                    n += 1
                continue
            cols = {_norm(str(h)): i for i, h in enumerate(header) if h}
            start = 1 if cols else 0
            for r in rows[start:]:
                if not r or all(c is None for c in r):
                    continue
                zeile = {str(k): (r[i] if i < len(r) else None) for k, i in cols.items()}
                z = _spalten_merken(zeile)
                if not z["bezeichnung"]:
                    continue
                fp.zeile_speichern(z)
                n += 1
    return n


# --- SQLite / Datenbank (ERP-Warenwirtschaft) ----------------------------
def import_sqlite(pfad: str, tabelle: str = None) -> int:
    """Liest Preise aus einer beliebigen SQLite-Datenbank (ERP/Warenwirtschaft).

    Sucht automatisch die passende Tabelle (bezeichnung + preis), falls keine
    angegeben. Gibt Anzahl uebernommener Preise zurueck.
    """
    import sqlite3
    conn = sqlite3.connect(pfad)
    try:
        cur = conn.cursor()
        cur.execute("SELECT name FROM sqlite_master WHERE type='table'")
        tabellen = [r[0] for r in cur.fetchall()]
        if not tabellen:
            return 0
        if tabelle and tabelle in tabellen:
            zieltab = tabelle
        else:
            # erste tabelle mit 'bezeichnung'/'text' UND einem preis-feld waehlen
            zieltab = None
            for t in tabellen:
                cols = [c[1].lower() for c in cur.execute(f"PRAGMA table_info('{t}')")]
                if any(k in " ".join(cols) for k in ("bezeichnung", "text", "artikel", "leistung")) \
                        and any(k in " ".join(cols) for k in ("ep", "preis", "chf", "betrag")):
                    zieltab = t
                    break
            if not zieltab:
                zieltab = tabellen[0]
        cur.execute(f"SELECT * FROM '{zieltab}'")
        spalten = [d[0] for d in cur.description]
        n = 0
        for row in cur.fetchall():
            zeile = {spalten[i]: row[i] for i in range(len(spalten))}
            z = _spalten_merken(zeile)
            if not z["bezeichnung"]:
                continue
            fp.zeile_speichern(z)
            n += 1
        return n
    finally:
        conn.close()


# --- Cloud (Google Sheets / OneDrive / SharePoint) via CSV-URL ----------
def import_url(url: str) -> int:
    """Laed eine veroeffentlichte Preisliste (CSV) von einer Cloud-URL.

    Funktioniert mit:
      - Google Sheets: 'CSV export'-Link (.../gviz/tq?tqx=out:csv oder /export?format=csv)
      - OneDrive/SharePoint: ' Embedd/Download als CSV '-Link
    Kein Login noetig (nur oeffentlich geteilte CSV-Links).
    """
    import urllib.request
    req = urllib.request.Request(url, headers={"User-Agent": "DevisPro/1.0"})
    with urllib.request.urlopen(req, timeout=30) as resp:
        data = resp.read()
    text = data.decode("utf-8-sig", "replace")
    import io
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return 0
    header = rows[0]
    cols = {_norm(str(h)): i for i, h in enumerate(header) if h}
    n = 0
    for r in rows[1:]:
        if not r or all(c is None or str(c).strip() == "" for c in r):
            continue
        zeile = {k: (r[i] if i < len(r) else None) for k, i in cols.items()}
        if not cols:  # keine erkennbaren spalten -> positionsbasiert
            zeile = {("bezeichnung" if i == 0 else f"col{i}"): (r[i] if i < len(r) else None)
                     for i in range(len(r))}
        z = _spalten_merken(zeile)
        if not z["bezeichnung"]:
            continue
        fp.zeile_speichern(z)
        n += 1
    return n


def import_any(pfad: str) -> int:
    """Dispatch nach Dateiendung: xlsx / pdf / sqlite / db / csv / txt."""
    ext = os.path.splitext(pfad)[1].lower()
    if ext in (".xlsx", ".xls"):
        return import_xlsx(pfad)
    if ext == ".pdf":
        return import_pdf(pfad)
    if ext in (".sqlite", ".db", ".sqlite3"):
        return import_sqlite(pfad)
    # csv / txt
    return fp.speichern_aus_upload(pfad)
