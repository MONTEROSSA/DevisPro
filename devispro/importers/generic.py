"""Generischer CSV/Excel-Importer fuer CH-KMU mit eigenen Spaltenlayouts.

Erkennt Spalten flexibel ueber Schlagworte (Position, Text, Menge,
Einheit, Preis, Betrag), damit beliebige Bau-Devis importiert werden
koennen - nicht nur Sorba.
"""
import csv
import io
from . import BaseImporter, register
from ..models import Devis, Position

# Spalten-Mustering: (Feld, Liste von Schlagworten im Header)
# DE + EN Keywords, damit beliebige Bau-Devis importiert werden koennen.
# Reihenfolge wichtig: spezifischere Felder (ep/betrag) VOR einheit pruefen,
# damit "unit_price" nicht fälschlich als einheit gemappt wird.
_PATTERNS = {
    "pos_nr": ("position", "pos", "nr", "nummer", "lfd", "no", "item", "artikelnr"),
    "text": ("text", "bezeichnung", "beschrieb", "leistung", "positionstext", "objekt",
             "description", "desc", "article", "name", "bezeich"),
    "menge": ("menge", "quant", "qty", "anzahl", "meng", "quantity", "amount"),
    "ep": ("einheitspreis", "unitprice", "unit_price", "ep", "preis", "ansatz",
           "preis/chf", "price", "rate"),
    "betrag": ("betrag", "total", "summe", "kosten", "gesamt", "linetotal",
               "lineTotal", "totalprice", "amount"),
    "einheit": ("einheit", "eh", "mass", "uom", "me"),
}


def _match_field(header: str):
    h = header.strip().lower()
    for field, kws in _PATTERNS.items():
        for kw in kws:
            if kw in h:
                return field
    return None


@register
class GenericCsvImporter(BaseImporter):
    name = "Generisches CSV/Excel"
    extensions = ("csv", "txt", "xlsx", "xls")

    def _rows(self, path):
        ext = path.lower().split(".")[-1]
        if ext in ("xlsx", "xls"):
            from openpyxl import load_workbook
            wb = load_workbook(path, data_only=True, read_only=True)
            ws = wb.active
            rows = []
            for r in ws.iter_rows(values_only=True):
                rows.append(["" if c is None else str(c) for c in r])
            return rows
        # CSV / TSV / semikolon-getrennt
        with open(path, encoding="utf-8-sig", errors="replace", newline="") as f:
            sample = f.read(4096)
            f.seek(0)
            # haeufigster Trenner unter tab / semikolon / komma
            cands = {"\t": sample.count("\t"), ";": sample.count(";"),
                     ",": sample.count(",")}
            delim = max(cands, key=cands.get)
            rd = csv.reader(f, delimiter=delim)
            return [r for r in rd]

    def parse(self, path: str) -> Devis:
        rows = self._rows(path)
        if not rows:
            return self._devis("(leer)", [])

        # Header-Zeile finden (erste Zeile mit erkanntem Feld)
        header_idx = 0
        colmap = {}
        for i, row in enumerate(rows[:5]):
            cmap = {}
            for ci, cell in enumerate(row):
                fld = _match_field(cell)
                if fld and fld not in cmap.values():
                    cmap[ci] = fld
            if "text" in cmap.values() or "menge" in cmap.values():
                header_idx = i
                colmap = cmap
                break
        if not colmap:
            # keine Header erkannt -> Spalten als pos,text,menge,einheit,ep annehmen
            colmap = {0: "pos_nr", 1: "text", 2: "menge", 3: "einheit", 4: "ep"}

        positions = []
        for row in rows[header_idx + 1:]:
            if not any(c.strip() for c in row):
                continue
            d = {}
            for ci, fld in colmap.items():
                d[fld] = row[ci] if ci < len(row) else ""
            text = d.get("text", "")
            if not text.strip():
                continue
            p = Position(
                pos_nr=d.get("pos_nr", "") or str(len(positions) + 1),
                text=text,
                menge=float(self._num(d.get("menge", 0)) or 0),
                einheit=d.get("einheit", ""),
                ep=(float(self._num(d["ep"])) if d.get("ep") else None),
                betrag=(float(self._num(d["betrag"])) if d.get("betrag") else None),
            )
            p.fill()
            positions.append(p)

        return self._devis("Import", positions)

    @staticmethod
    def _num(s):
        if s is None:
            return 0
        s = str(s).strip().replace("'", "").replace(" ", "")
        # CHF-Trennzeichen
        s = s.replace("’", "").replace("’", "")
        for sep in ("'", " "):
            s = s.replace(sep, "")
        s = s.replace("'", "")
        # apostroph als tausender
        if s.count(".") > 1:
            s = s.replace(".", "")
        try:
            return float(s)
        except ValueError:
            return 0
